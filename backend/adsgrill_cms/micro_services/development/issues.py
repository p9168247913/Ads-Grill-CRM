from app.models import Users,Project,Sprint,Issue,LinkedIssue,WorkLog
from rest_framework.views import APIView
from django.http import JsonResponse
from django.db.models import ObjectDoesNotExist
from django.db import transaction
import os
from django.utils import timezone
from rest_framework import status
import zipfile
from django.http import HttpResponse
from io import BytesIO
import uuid
from django.db.models import Sum
from datetime import timedelta
import re
from rest_framework.authentication import SessionAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import AuthenticationFailed
from braces.views import CsrfExemptMixin
from django.conf import settings
from django.core.mail import send_mail
import asyncio
from datetime import datetime, time
import threading
import pandas as pd
import openpyxl
import json
import uuid


class CsrfExemptSessionAuthentication(SessionAuthentication):
    def enforce_csrf(self, request):
        return
    
class CustomSessionAuthentication(SessionAuthentication):
    def authenticate(self, request):
        user_auth_tuple = super().authenticate(request)
        if user_auth_tuple is None:
            raise AuthenticationFailed('Your session has expired. Please log in again.')
        return user_auth_tuple

def convert_to_duration(duration_str):
    if 'd' not in duration_str:
        duration_str+=' 0d'
    if 'h' not in duration_str:
        duration_str+=' 0h'
    if 'm' not in duration_str:
        duration_str+=' 0m'
    regex = re.compile(r'(?:(?P<weeks>\d+)w )?(?:(?P<days>\d+)d )?(?:(?P<hours>\d+)h )?(?:(?P<minutes>\d+)m)?')

    match = regex.match(duration_str)
    components = {key: int(value) if value else 0 for key, value in match.groupdict().items()}

    duration = timedelta(
        weeks=components.get('weeks', 0),
        days=components.get('days', 0),
        hours=components.get('hours', 0),
        minutes=components.get('minutes', 0)
    )
    return duration

class IssueView(CsrfExemptMixin, APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, CustomSessionAuthentication]
    permission_classes = [IsAuthenticated]

    def send_issue_details(self,title, issue_type,priority,created_at,reporter,link,email):
        local_time=created_at.astimezone(timezone.get_current_timezone())
        time=local_time.strftime('%d-%m-%y  %H:%M:%S')
        subject = f'{title}'
        message = f'Created_By : {reporter}\n \n Type : {issue_type} \n \n Priority : {priority} \n \n For More Details : {link} \n \n Created_at : {time}'
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient = [email]
        send_mail(subject,message,from_email, recipient)
    
    def post(self, request):
        try:
            requestData = request.data 
            project_id = requestData.get('project_id')
            sprint_id = requestData.get('sprint_id')
            reporter_id = requestData.get('reporter_id')    
            team_lead_id = requestData.get('team_lead_id')
            title = requestData.get('title')
            key = requestData.get('key')
            description = requestData.get('description')
            type = requestData.get('type')
            priority = requestData.get('priority')
            attachments = request.FILES.getlist('attachments', [])
            assignee_id=requestData.get('assignee_id')
            parent_issues=requestData.getlist('parent_issues',[])
            exp_duration = requestData.get('exp_duration')
            linked_issues=requestData.getlist('linked_issues',[])
            linked_issue_type=requestData.getlist('linked_issue_type')
            
            sprint_instance = Sprint.objects.get(pk=sprint_id)
            sprint_exp_duration=sprint_instance.exp_duration

            if Issue.objects.filter(title=title, project=project_id, sprint = sprint_id).exists():
                return JsonResponse({"message": "Issue with this title already exists"})
                
            issue_exp_duration = convert_to_duration(exp_duration)
            existing_issue_duration=Issue.objects.filter(sprint=sprint_instance).aggregate(Sum('exp_duration'))['exp_duration__sum']
            if existing_issue_duration:
                existing_issue_duration = existing_issue_duration
            else:
                existing_issue_duration = timedelta(0)
            
            total_duration = issue_exp_duration + existing_issue_duration

            if total_duration > sprint_exp_duration:
                return JsonResponse({"message": "Creating this issue will affect the active sprint's scope, please update sprint duration"})

            with transaction.atomic():
                project_instance = Project.objects.get(pk=project_id)
                reporter_instance = Users.objects.get(pk=reporter_id)
                team_lead_instance = Users.objects.get(pk=team_lead_id) if team_lead_id else None
                assignee_instance=Users.objects.get(pk=assignee_id)
                
                issue_instance = Issue.objects.create(
                    project=project_instance,
                    sprint=sprint_instance,
                    reporter=reporter_instance,
                    team_lead=team_lead_instance,
                    assignee=assignee_instance,
                    title=title,
                    key = key,
                    description=description,
                    type=type,
                    priority=priority,
                    exp_duration=exp_duration
                )
              
                attachment_file_names = []
                for attachment in attachments:
                    subdirectory = os.path.join('media', 'uploads', 'Development', 'issues', str(f"{project_instance.key}_{sprint_instance.key}_{issue_instance.title}"))
                    if not os.path.exists(subdirectory):
                        os.makedirs(subdirectory)
                        
                    file_extension = attachment.name.split('.')[-1]
                    unique_id = str(uuid.uuid4().hex[:6])
                    unique_filename = f"{unique_id}_{issue_instance.key}.{file_extension}"
                    attachment_path = os.path.join(subdirectory, unique_filename)
                    with open(attachment_path, 'wb') as destination:
                        for chunk in attachment.chunks():
                            destination.write(chunk)

                    attachment_file_names.append(attachment_path)

                issue_instance.attachments = attachment_file_names
                issue_instance.save()
                
                if parent_issues:
                    parent_issues = Issue.objects.filter(pk__in=parent_issues)
                    if not parent_issues:
                        return JsonResponse({"message":"Parent issues does not exists or invalid request"}, status=status.HTTP_400_BAD_REQUEST)
                    issue_instance.parent_issue.add(*parent_issues)
                    issue_instance.save()

                if linked_issues:                           
                    linked_issues = Issue.objects.filter(pk__in=linked_issues)
                    if not linked_issues:
                        return JsonResponse({"message":"Linked issues does not exists or invalid request"}, status=status.HTTP_400_BAD_REQUEST)
                    linked_issue_instances = [
                        LinkedIssue(
                            project=project_instance,
                            sprint=sprint_instance,
                            destination=issue_instance,
                            type=linked_issue_type,
                            source=linked_issue
                        )
                        for linked_issue in linked_issues]

                    LinkedIssue.objects.bulk_create(linked_issue_instances)
                                       
               # url = f'http://127.0.0.1:8000/api/development/issues?id={issue_instance.pk}'
                url = 'https://crm.adsgrill.com/issues'
                send_email = threading.Thread(target=self.send_issue_details, args=((issue_instance.title, issue_instance.type, issue_instance.priority, issue_instance.created_at, issue_instance.reporter.name, url, issue_instance.assignee.email)))
                send_email.start()
        
        except Project.DoesNotExist:
            return JsonResponse({"message": "Requested project does not exists"})

        except Sprint.DoesNotExist:
            return JsonResponse({"message": "Requested sprint does not exists"})

        except Users.DoesNotExist:
            return JsonResponse({"message": "Requested user does not exists"})
        except Exception as e:
            return JsonResponse(str(e))
        return JsonResponse({"message": "Issue created successfully"},status=status.HTTP_201_CREATED)
    
    def get(self,request):
        try:
            project_instance=Project.objects.get(pk=request.GET.get('project_id'))
            all_issues=Issue.objects.filter(project=project_instance).order_by('-created_at')
            if not all_issues.exists():
                return JsonResponse({"message":"No issues on this project"}, status=status.HTTP_204_NO_CONTENT)
            issue_data = [{
                'id':issue.pk,
                "project": issue.project.name,
                "sprint": {
                    'id':issue.sprint.pk,
                    'name':issue.sprint.name
                },
                "reporter": {
                    "id":issue.reporter.pk,
                    "name":issue.reporter.name
                },
                "team_lead": {
                    "id": issue.team_lead.pk if issue.team_lead else None,
                    "name": issue.team_lead.name if issue.team_lead else None
                },
                "title": issue.title,
                "key": issue.key,
                "description": issue.description,
                "type": issue.type,
                "priority": issue.priority,
                "status": issue.status,
                "attachments": issue.attachments,
                "exp_duration": issue.exp_duration,
                "org_duration": issue.org_duration,
                "created_at": issue.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                "assignee": {
                    "id":issue.assignee.pk,
                    "name":issue.assignee.name
                }
            } for issue in all_issues]
                
        except Project.DoesNotExist:
            return JsonResponse({"message":"Invalid Credentials(project)"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({str(e)})    
        
        return JsonResponse({"issues":issue_data},status=status.HTTP_200_OK)
    
    def put(self,request):
        try:
            requestData=request.data
            upd_issue=Issue.objects.get(pk=requestData.get('id'))
            project_instance=Project.objects.get(pk=requestData.get('project_id'))
            sprint_instance=Sprint.objects.get(pk=requestData.get('sprint_id'))
            reporter_instance=Users.objects.get(pk=requestData.get('reporter_id'))
            team_lead_instance=Users.objects.get(pk=requestData.get("team_lead_id")) if requestData.get('team_lead_id') else None
            assignee_instance=Users.objects.get(pk=requestData.get("assignee_id"))
            parent_issues=requestData.getlist('parent_issues',[])
            attachments = request.FILES.getlist('attachments', [])
            
            sprint_exp_duration=sprint_instance.exp_duration
            issue_exp_duration = convert_to_duration(request.data.get('exp_duration'))
            existing_issue_duration=Issue.objects.filter(sprint=sprint_instance).exclude(pk=upd_issue.pk).aggregate(Sum('exp_duration'))['exp_duration__sum']
            if existing_issue_duration:
                existing_issue_duration = existing_issue_duration
            else:
                existing_issue_duration = timedelta(0)
            
            total_duration = issue_exp_duration + existing_issue_duration

            if total_duration > sprint_exp_duration:
                return JsonResponse({"message": "Creating this issue will affect the active sprint's scope, please update sprint duration"})
            
            
            if(request.user.designation=="project_manager" and request.user.role.name=="development"):
                if requestData.get("status")=="done":
                    worklogs=WorkLog.objects.filter(issue=upd_issue).order_by('-created_at')
                    if worklogs:
                        total_org_duration=worklogs.aggregate(Sum('logged_time'))['logged_time__sum']
                        upd_issue.org_duration=total_org_duration
                    else:
                        upd_issue.org_duration=None
                        return JsonResponse({"message":"No worklogs found for this issue, can't change status to done"})
                    
            with transaction.atomic():
                upd_issue.sprint=sprint_instance
                upd_issue.project=project_instance
                upd_issue.reporter=reporter_instance
                upd_issue.team_lead=team_lead_instance
                upd_issue.assignee=assignee_instance
                upd_issue.status=requestData.get("status")
                upd_issue.title=requestData.get('title')
                upd_issue.key=requestData.get('key')
                upd_issue.description=requestData.get('description')
                upd_issue.type=requestData.get('type')
                upd_issue.priority=requestData.get('priority')
                upd_issue.exp_duration=requestData.get('exp_duration')
                if parent_issues:
                    upd_issue.parent_issue.clear()     
                    parent_issues = Issue.objects.filter(pk__in=parent_issues)
                    upd_issue.parent_issue.add(*parent_issues)

                attachment_file_names = []
                for attachment in attachments:
                    subdirectory = os.path.join('media', 'uploads', 'Development', 'issues', str(f"{project_instance.key}_{sprint_instance.key}_{upd_issue.key}"))
                    if not os.path.exists(subdirectory):
                        os.makedirs(subdirectory)
                    file_extension = attachment.name.split('.')[-1]
                    unique_id = str(uuid.uuid4().hex[:6])
                    unique_filename = f"{unique_id}_{upd_issue.key}.{file_extension}"
                    attachment_path = os.path.join(subdirectory, unique_filename)
                    with open(attachment_path, 'wb') as destination:
                        for chunk in attachment.chunks():
                            destination.write(chunk)

                    attachment_file_names.append(attachment_path)
                upd_issue.attachments.extend(attachment_file_names)
                upd_issue.save()
                
        except Project.DoesNotExist:
            return JsonResponse({"message":"Requested Project not exists"})
        except Sprint.DoesNotExist:
            return JsonResponse({"message":"Requested Sprint not exists"})
        except Users.DoesNotExist:
            return JsonResponse({"message":"Requested User not exists"})
        except Exception as e:
            return JsonResponse({"message":str(e)})
        return JsonResponse({"message":"issue Updated successfully"})
    
    def delete(self,request):
        try:
            issue_instance=Issue.objects.get(pk=request.GET.get('id'))
            file_paths = issue_instance.attachments
            if file_paths:
                for file_path in file_paths:
                    if os.path.exists(file_path):
                        os.remove(file_path)
            directory = os.path.join('media', 'uploads', 'Development', 'issues', str(f"{issue_instance.project.key}_{issue_instance.sprint.key}_{issue_instance.key}"))
            directory_path = os.path.join("media\\uploads\\Development\\projects\\", directory)
            if os.path.exists(directory_path):
                os.rmdir(directory_path)
            issue_instance.delete()
        except Issue.DoesNotExist:
            return JsonResponse({"message":"Request Issue not exists"})
        return JsonResponse({"message":"issue deleted successfully"})
    
class LinkedIssueView(CsrfExemptMixin, APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]
    def post(self,request):
        try:
            requestData=request.data
            source_issues=requestData.get("source_issues")
            destination_issue=requestData.get("destination_issue")
            sourceType=requestData.get("type")
            destination_instance=Issue.objects.get(pk=destination_issue)
            
            if source_issues:
                source_issues_list=source_issues.split(',')
                check_exists=LinkedIssue.objects.filter(
                    source__in=source_issues_list,
                    destination=destination_instance
                ).exists()

                if check_exists:
                    return JsonResponse({"message":"issue already mapped"})
                else:
                    issues = Issue.objects.filter(pk__in=source_issues_list)
                    linked_issue_instances=[
                        LinkedIssue(
                        project=source.project,
                        sprint=source.sprint,
                        destination=destination_instance,
                        source = source,
                        type=sourceType
                    ) for source in issues]

                    LinkedIssue.objects.bulk_create(linked_issue_instances)
            else:
                return JsonResponse({'message':'Source issues not found'}, status=status.HTTP.HTTP_400_BAD_REQUEST)
        except Project.DoesNotExist:
            return JsonResponse({"message":"Requested Project not exists"})
        
        except Sprint.DoesNotExist:
            return JsonResponse({"message":"Requested Sprint not exists"})
        
        except Issue.DoesNotExist:
            return JsonResponse({"message":"Requested Issue not exists"})
        
        except Exception as e:
            return JsonResponse({"errror":str(e)})
        
        return JsonResponse({"message":"Issues linked successfully "})
    
class DownloadIssuesAttchments(CsrfExemptMixin, APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, CustomSessionAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self,request):
        try:
            id=request.GET.get("id")
            issue_instance=Issue.objects.get(pk=id)
            files=issue_instance.attachments
            if files:
                zip_buffer=BytesIO()
                with zipfile.ZipFile(zip_buffer,'w') as pro_zip:
                    for file in files:
                        file_name = file.split("\\")[-1]
                        pro_zip.write(file, file_name)
                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename={issue_instance.key}_attachments.zip'
            else:
                response = JsonResponse({"message":"No files found for the specified criteria."}, status=status.HTTP_204_NO_CONTENT)
            return response
            
        except Exception as e:
            return JsonResponse({"message":str(e)})
        
class IssueMetaData(APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, CustomSessionAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self,request):
        try:
            id=request.GET.get('id')
            issue=Issue.objects.get(pk=id)

            issue_data={
                "id":issue.pk,
                "title":issue.title,
                "key":issue.key,
                "description":issue.description,
                "priority":issue.priority,
                "status":issue.status,
                "attachments":issue.attachments,
                "exp_duration":issue.exp_duration,
                "org_duration":issue.org_duration,
                "created_at":issue.created_at,
                "sprint":{
                    "id":issue.sprint.pk,
                    "key":issue.sprint.key,
                    "name":issue.sprint.name
                },
                "project":{
                    "id":issue.project.pk,
                    "key":issue.project.key,
                    "name":issue.project.name,
                },
                "reporter":{
                    "id":issue.reporter.pk,
                    "email":issue.reporter.email,
                    "role":issue.reporter.role.name,
                    "designation":issue.reporter.designation,
                },
               
            }
            if issue.assignee is not None:
                issue_data['assignee']={
                    "id":issue.assignee.pk,
                    "email":issue.assignee.email,
                    "role":issue.assignee.role.name,
                    "designation":issue.assignee.designation,
                }
            else : issue_data['assignee']={}
            
            if issue.team_lead is not None:
                issue_data["team_lead"]={
                    "id":issue.team_lead.pk,
                    "email":issue.team_lead.email,
                    "role":issue.team_lead.role.name,
                    "designation":issue.team_lead.designation,
                }
            else : issue_data['team_lead']={}
            
            parent_issues = Issue.objects.filter(pk__in=issue.parent_issue.values_list('pk')) 
            if parent_issues.exists():
                issue_data['parentIssues']=[{
                    "id":parent_issue.pk,
                    "key":parent_issue.key,
                    "title":parent_issue.title,
                    'type':parent_issue.type,
                    "status":parent_issue.status,
                    "priority":parent_issue.priority,
                    "assignee":parent_issue.assignee.name
                }for parent_issue in parent_issues]
            else : issue_data['parentIssues']=[]
                
                
            child_issues=Issue.objects.filter(parent_issue__in=[id])
            if child_issues.exists():
                issue_data['childIssues']=[{
                    "id":child_issue.pk,
                    "key":child_issue.key,
                    "title":child_issue.title,
                    'type':child_issue.type,
                    "status":child_issue.status,
                    "priority":child_issue.priority,
                    "assignee":child_issue.assignee.name
                } for child_issue in child_issues]
            else : issue_data["childIssues"]=[]
                
            linked_issues = LinkedIssue.objects.filter(destination__in=[id])
            if linked_issues:
                issue_data['linked_issues']=[{
                    "id":linked_issue.source.pk,
                    "link_type":linked_issue.type,
                    "key":linked_issue.source.key,
                    "title":linked_issue.source.title,
                    "type":linked_issue.source.type,
                    "status":linked_issue.source.status,
                    "priority":linked_issue.source.priority,
                    "assignee":linked_issue.source.assignee.name
                }for linked_issue in linked_issues]
            else: issue_data['linked_issues'] = []
                            
        except Issue.DoesNotExist:
            return JsonResponse({'message':'Requested issue not exists'})
            
        except Exception as e:
            return JsonResponse({'error':str(e)})
        
        return JsonResponse({"Data":issue_data},status=status.HTTP_200_OK)
    
class downloadUserWorkReport(APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, CustomSessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        id = request.GET.get("id")
        date_range=request.GET.get('date_range') if request.GET.get('date_range') else None
        if not id:
            return JsonResponse({"message": "Please provide Employee id"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            
            if date_range is not None:
                date_range = json.loads(date_range)
                startDate_str = date_range['start_date']
                endDate_str = date_range['end_date']
                start_datetime = datetime.strptime(startDate_str, "%Y-%m-%d").date()
                end_datetime = datetime.strptime(endDate_str, "%Y-%m-%d").date()
            
            user_instance = Users.objects.get(pk=id)
            all_issues = Issue.objects.filter(assignee=user_instance).order_by("-created_at")
            
            if date_range is not None: 
                all_issues = all_issues.filter(created_at__date__gte=start_datetime,created_at__date__lte=end_datetime).order_by('-created_at') 
            
            wb=openpyxl.Workbook()
            ws=wb.active
            
            ws['A1'] = 'Assignee'
            ws['B1'] = 'Designation'
            ws['C1'] = 'Project Name'
            ws['D1'] = 'Sprint Name'
            ws['E1'] = 'Issue Name'
            ws['F1'] = 'Status'
            ws['G1'] = 'Expected Estimate'
            ws['H1'] = 'Original Estimate'
            ws['I1'] = 'Logged Time'
            ws['J1'] = 'Remaining Time'
            ws['K1'] = 'Log Created At'
            ws['L1'] = 'Total Logged Time'
            ws['M1'] = 'Extra Effort'
            # ws['N1'] = 'Comments'
            
            
            for issue in all_issues:
                all_worklogs=WorkLog.objects.filter(issue=issue).order_by('-created_at')
                total_logged_time=all_worklogs.aggregate(Sum('logged_time'))['logged_time__sum']
                total_extra_effort=all_worklogs.aggregate(Sum('extra_efforts'))['extra_efforts__sum']
                for worklog in all_worklogs:
                    ws.append([
                        worklog.issue.assignee.name,
                        worklog.issue.assignee.designation,
                        worklog.issue.sprint.project.name,
                        worklog.issue.sprint.name,
                        worklog.issue.title,
                        worklog.issue.status,
                        worklog.issue.exp_duration,      
                        worklog.issue.org_duration if worklog.issue.org_duration else "N/A",
                        worklog.logged_time,
                        worklog.remaining_time,
                        worklog.created_at.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S'),
                        total_logged_time,
                        total_extra_effort,
                        # worklog.description
                    ])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{user_instance.name}_report.xlsx.xlsx"'
            response.write(output.getvalue())
        
        except Issue.DoesNotExist:
            return JsonResponse({"message":"Issues on this user not exists"},status=status.HTTP_404_NOT_FOUND)
        except Users.DoesNotExist:
            return JsonResponse({"message": "Requested user not exist"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return JsonResponse({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return response
                
        
        
    
