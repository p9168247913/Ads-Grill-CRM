from app.models import Users, Lead, Source, Tag, LeadStatus,Sale
from braces.views import CsrfExemptMixin
from rest_framework.views import APIView
from django.http import JsonResponse, HttpResponse
from django.db.models import ObjectDoesNotExist
from django.db import transaction
import pandas as pd
import json
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, FileUploadParser
from django.db.utils import IntegrityError
from os.path import basename
import os
import openpyxl
from rest_framework.authentication import SessionAuthentication
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime,timedelta


class CsrfExemptSessionAuthentication(SessionAuthentication):
    def enforce_csrf(self, request):
        return

class CustomSessionAuthentication(SessionAuthentication):
    def authenticate(self, request):
        user_auth_tuple = super().authenticate(request)
        if user_auth_tuple is None:
            raise AuthenticationFailed('Your session has expired. Please log in again.')
        return user_auth_tuple

class LeadView(CsrfExemptMixin, APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, CustomSessionAuthentication]
    permission_classes = [IsAuthenticated]
    def post(self, request):
        parser_classes = (MultiPartParser, FormParser)
        lead_data = request.data
        val = lead_data.get('key')
        if not lead_data or not val:
            return JsonResponse({'message':'Bad Request'}, status=status.HTTP_400_BAD_REQUEST)
        
        user_instance=Users.objects.get(email=request.user.email)
        if user_instance.role.name != "leads":
            return JsonResponse({'message':"Sorry you can not add lead"},status=status.HTTP_401_UNAUTHORIZED)
        
        lead_instance = Lead.objects.filter(Q(contact_no=lead_data.get('contact_no')) | Q(email=lead_data.get('email')))
        
        if lead_instance.exists():
            return JsonResponse({'message':'Lead with this contact or email already exists'},status=status.HTTP_400_BAD_REQUEST)
        
        if val =='post':
            try:
                with transaction.atomic():            
                    source_instance, _ = Source.objects.get_or_create(name=lead_data.get('source'))
                    
                    if lead_data.get('contact_no')==None:
                        return JsonResponse({'message':"Please Enter Contact Details"},status=status.HTTP_400_BAD_REQUEST)

                    # Create Lead instance with related models
                    create_lead = Lead.objects.create(
                        sales_man=user_instance,
                        source=source_instance,
                        client_name=lead_data.get('client_name'),
                        contact_no=lead_data.get('contact_no'),
                        email=lead_data.get('email'),
                        requirement=lead_data.get('requirement')
                    )
                    create_lead.save()
            except IntegrityError as i:
                return JsonResponse({'message':str(i)}, status=status.HTTP_400_BAD_REQUEST)
            return JsonResponse({'message':'Lead Created Successfully'},status=status.HTTP_201_CREATED)
        
        if val == 'bulkUpload':
            
            parser_class = (FileUploadParser,)
            up_file = request.FILES.get('file')
            upload_folder = 'uploads/leads'
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
        
            destination_path = os.path.join(upload_folder, up_file.name)
            with open(destination_path, "wb") as destination:
                for i, chunk in enumerate(up_file.chunks()):
                    destination.write(chunk)
        
            df = pd.read_excel(destination_path, dtype=str)
            original_len = len(df)
            df = df.drop_duplicates(subset=['Contact No'])
            new_len = len(df)
            if new_len < original_len:
                excel_response = {"message":f"{original_len-new_len} Duplicates(contact no) were founded and deleted"}
            else:
                excel_response = {"message":"No duplicates found in the excel file"}

            existing_rows = Lead.objects.values_list('contact_no', flat=True)
            df = df[~df['Contact No'].isin(existing_rows)]
            if df.empty:
                database_response = {"message":"No new records to save to the database"}
            else:
                df = df.dropna(how='all')
                empty_cols = []
            
                try:
                    with transaction.atomic():
                        for ind, row in df.iterrows():
                            xl_source, xl_cl_name, xl_email, xl_contact_no, xl_requirement = row
                            xl_cl_name = str(xl_cl_name)
            
                            if pd.isna(xl_contact_no) or not str(xl_contact_no):
                                empty_cols.append('contact_no')
            
                            if empty_cols:
                                col_name = ', '.join(empty_cols)
                                raise Exception(f'Empty fields in columns: {col_name}, Row {ind+2}')
                            else:
                                source_instance, _ = Source.objects.get_or_create(name=xl_source)
            
                                create_lead = Lead.objects.create(
                                    sales_man=user_instance,
                                    source=source_instance,
                                    client_name=xl_cl_name,
                                    contact_no=xl_contact_no,
                                    email=xl_email,
                                    requirement=xl_requirement
                                )
                                create_lead.save()
                        database_response = {"message":f"{up_file.name} uploaded successfully", "status":status.HTTP_201_CREATED}
        
                except IntegrityError as i:
                    return JsonResponse({'message': str(i)}, status=status.HTTP_400_BAD_REQUEST)
                
                except ObjectDoesNotExist:
                    return JsonResponse({'message': f"No Sales Manager Found At Row: {ind+2}"},status=status.HTTP_404_NOT_FOUND)
            
                except Exception as exc:
                    return JsonResponse({'message': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
            
            return JsonResponse({"excel_res":excel_response, "database_res":database_response})

    
    def get(self, request):
        try:
            pageNo = request.GET.get("page_no")
            client_name = request.GET.get('client_name') if request.GET.get('client_name') else None
            contact_no = request.GET.get('contact_no') if request.GET.get('contact_no') else None
            date_range=request.GET.get('date_range') if request.GET.get('date_range') else None
            noOfRecords = 15
            
            if date_range is not None:
                date_range = json.loads(date_range)
                startDate_str = date_range['start_date']
                endDate_str = date_range['end_date']
                start_datetime = datetime.strptime(startDate_str, "%Y-%m-%d").date()
                end_datetime = datetime.strptime(endDate_str, "%Y-%m-%d").date()
                            
            with transaction.atomic():
                allLeads = Lead.objects.filter(is_deleted=False).order_by('-created_at')
                if client_name is not None:
                    allLeads = allLeads.filter(client_name__icontains=client_name,is_deleted=False).order_by('-created_at')

                if contact_no is not None:
                    allLeads = allLeads.filter(contact_no__icontains=contact_no,is_deleted=False).order_by('-created_at') 
                   
                if date_range is not None: 
                    allLeads = allLeads.filter(created_at__date__gte=start_datetime,created_at__date__lte=end_datetime,is_deleted=False).order_by('-created_at') 
                    
                if client_name is not None and contact_no is not None:
                    allLeads = allLeads.filter(Q(client_name__icontains=client_name) | Q(contact_no__icontains=contact_no),is_deleted=False).order_by('-created_at')

                p = Paginator(allLeads, noOfRecords)
                page_obj = p.get_page(pageNo)
                
                try:
                    page_obj = p.page(pageNo)
                except PageNotAnInteger:
                    page_obj = p.page(1)
                except EmptyPage:
                    page_obj = p.page(p.num_pages)
                
                lead_count = allLeads.count()
                assignedLeads = Sale.objects.all().count()
                unassignedLeads = lead_count - assignedLeads
                
                leads_data = [{
                    'id': lead.pk,
                    'source': lead.source.name,
                    'client_name': lead.client_name,
                    'email': lead.email,
                    'contact_no': lead.contact_no,
                    'requirement': lead.requirement,
                    'is_assigned':lead.is_assigned,
                    'date': lead.created_at
                } for lead in page_obj]
                
                data = {
                    'leads': leads_data,
                    'total_leads': lead_count,
                    'assigned_leads': assignedLeads,
                    'unassigned_leads': unassignedLeads,
                    'total_pages': p.num_pages
                }
                   
        except IntegrityError as i:
            return JsonResponse({'message': str(i)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'message':str(e)},status=status.HTTP_400_BAD_REQUEST)
        
        return JsonResponse({'lead_data': data}, status=status.HTTP_200_OK)
        
    def put(self, request):
        parser_classes = (MultiPartParser, FormParser)
        lead_data = request.data
        if not lead_data:
            return JsonResponse({'message':'Bad Request'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user_instance = Users.objects.get(email=request.user)            
            if user_instance.role.name != "sales" and user_instance.role.name != "leads":
                return JsonResponse({"message":"Sorry! You Can Not edit the lead"},status=status.HTTP_401_UNAUTHORIZED)
            
            if lead_data.get('contact_no')==None:
                return JsonResponse({'message':"Please Enter Contact No."},status=status.HTTP_400_BAD_REQUEST)
            
            with transaction.atomic():
                updateLead = Lead.objects.get(pk=lead_data.get('id'))
                source_instance, _ = Source.objects.get_or_create(name=lead_data.get('source'))
                
                updateLead.sales_man = user_instance
                updateLead.source =  source_instance
                updateLead.client_name =  lead_data.get('client_name')
                updateLead.email =  lead_data.get('email')
                updateLead.contact_no=lead_data.get('contact_no')
                updateLead.requirement=lead_data.get('requirement')
                updateLead.save()
        except IntegrityError as i:
            return JsonResponse({'message':str(i)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'message':str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return JsonResponse({'message':'Lead Updated Successfully'}, status=status.HTTP_200_OK)
        
        
    def delete(self,request):
        id = request.GET.get('id')
        if not id:
            return JsonResponse({'message':'Bad Request'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            with transaction.atomic():
                deleteLead = Lead.objects.get(pk=id)
                deleteLead.delete()
        except IntegrityError as i:
            return JsonResponse({'message':str(i)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'message':str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return JsonResponse({'message':'Lead deleted Successfully'}, status=status.HTTP_204_NO_CONTENT)
        
        
        
class LeadInfo(CsrfExemptMixin, APIView):
    authentication_classes = [CsrfExemptSessionAuthentication, CustomSessionAuthentication]
    permission_classes = [IsAuthenticated]
    def post(self, request):
        val = request.GET.get('key')
        name = request.GET.get('name')
        if not name or not val:
            return JsonResponse({'message':'Bad Request'}, status=status.HTTP_400_BAD_REQUEST)
        if val == 'status':
            try:
                with transaction.atomic():
                    leadStatusInstance = LeadStatus.objects.create(name=name)
                    leadStatusInstance.save()
                    return JsonResponse({'message':'Lead Status Created Successfully'}, status=status.HTTP_201_CREATED)
            except IntegrityError as i:
                return JsonResponse({'message':str(i)}, status=status.HTTP_400_BAD_REQUEST)
            
        if val == 'tag':
            try:
                with transaction.atomic():
                    leadTagInstance = Tag.objects.create(name=name)
                    leadTagInstance.save()
                    return JsonResponse({'message':'Lead Tag Created Successfully'}, status=status.HTTP_201_CREATED)
            except IntegrityError as i:
                return JsonResponse({'message':str(i)}, status=status.HTTP_400_BAD_REQUEST)
            
        if val == 'source':
            try: 
                with transaction.atomic():
                    leadSourceInstance = Source.objects.create(name=name)
                    leadSourceInstance.save()
                    return JsonResponse({'message':'Lead Source Created Successfully'}, status=status.HTTP_201_CREATED)
            except IntegrityError as i:
                return JsonResponse({'message':str(i)}, status=status.HTTP_400_BAD_REQUEST)
        
    def get(self, request):
        try:
            allLeadStatus = LeadStatus.objects.all().order_by('-created_at')
            allLeadTags = Tag.objects.all().order_by('-created_at')
            allLeadSource = Source.objects.all().order_by('-created_at')
            leadInfoData = {}

            leadStatus = [{
                'id':status.pk,
                'name':status.name
            } for status in allLeadStatus]

            leadTag = [{
                'id': tag.pk,
                'name': tag.name
            } for tag in allLeadTags]

            leadSource = [{
                'id': source.pk,
                'name': source.name
            } for source in allLeadSource]

            leadInfoData['leadStatus'] = leadStatus
            leadInfoData['leadTag'] = leadTag
            leadInfoData['leadSource'] = leadSource

            return JsonResponse({'leadInfoData':leadInfoData}, status=status.HTTP_200_OK)
        except Exception as e:
            return JsonResponse({'message':str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    def delete(self, request):
        val = request.GET.get('key')
        id = request.GET.get('id')
        if not id or not val:
            return JsonResponse({'message': 'Bad Request'}, status=status.HTTP_400_BAD_REQUEST)
        if val == 'status':
            try:
                with transaction.atomic():
                    deleteLeadStatus = LeadStatus.objects.get(pk=id)
                    deleteLeadStatus.delete()
            except ObjectDoesNotExist:
                return JsonResponse({'message': 'Lead Status Not Found'}, status=status.HTTP_404_NOT_FOUND)
            except IntegrityError as i:
                return JsonResponse({'message': str(i)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            return JsonResponse({'message': 'Lead Status Deleted Successfully'}, status=status.HTTP_200_OK)
            
        if val == 'tag':
            try:
                with transaction.atomic():
                    deleteLeadTag = Tag.objects.get(pk=id)
                    deleteLeadTag.delete()
            except ObjectDoesNotExist:
                return JsonResponse({'message': 'Lead Tag Not Found'}, status=status.HTTP_404_NOT_FOUND)
            except IntegrityError as i:
                return JsonResponse({'message': str(i)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            return JsonResponse({'message': 'Lead Tag Deleted Successfully'}, status=status.HTTP_200_OK)
            
        if val == 'source':
            try:
                with transaction.atomic():
                    deleteLeadSource = Source.objects.get(pk=id)
                    deleteLeadSource.delete()
            except ObjectDoesNotExist:
                return JsonResponse({'message': 'Lead Source Not Found'}, status=status.HTTP_404_NOT_FOUND)
            except IntegrityError as i:
                return JsonResponse({'message': str(i)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            return JsonResponse({'message': 'Lead Source Deleted Successfully'}, status=status.HTTP_200_OK)

class LeadExcelDownload(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Source'
        ws['B1'] = 'Client Name'
        ws['C1'] = 'Email'
        ws['D1'] = 'Contact No'
        ws['E1'] = 'Requirement'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        wb.save(response)
        return response 


